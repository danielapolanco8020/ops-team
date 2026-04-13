import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# Keywords that introduce a unit (Floor/Fl handled separately below)
UNIT_KEYWORDS = r'(?:Apt|Apartment|Unit|Suite|Ste|Room|Rm|Dept|Bldg|Building|Lot|#|No\.?|Num\.?)'

UNIT_PATTERNS = [
    # Ordinal floor: "2nd Floor", "3rd Fl"  — must come before keyword catch-all
    r'(?i)\b\d+(?:st|nd|rd|th)\s+(?:Floor|Fl)\b',
    # "Floor 3" / "Fl 2"
    r'(?i)\bFl(?:oor)?\s+\d+\b',
    # Keyword + alphanumeric: "Apt A11", "Unit 3B", "Suite 200", "Bldg C"
    rf'(?i)\b{UNIT_KEYWORDS}\.?\s*[A-Za-z0-9][A-Za-z0-9-]*',
    # Bare # sign + identifier: "#5", "#F5", "#4D"
    r'#\s*[A-Za-z0-9][A-Za-z0-9-]*',
    # Alphanumeric code after a comma: "123 Main St, 4D"
    r',\s*(?:[A-Za-z]\d|\d[A-Za-z])[A-Za-z0-9-]*\b',
    # Alphanumeric code after a dash: "123 Main St - B2"
    r'\s[-–]\s*(?:[A-Za-z]\d|\d[A-Za-z])[A-Za-z0-9-]*\b',
    # Trailing bare alphanumeric room code at end of string: "4D", "F5", "B12"
    r'(?i)\s+\b(?:[A-Za-z]\d{1,3}|\d{1,3}[A-Za-z])\s*$',
]


def find_unit(address: str):
    """Return (main_address, unit_string) for a given address string."""
    for pattern in UNIT_PATTERNS:
        match = re.search(pattern, address)
        if match:
            main = address[:match.start()].strip().rstrip(',').strip()
            unit = address[match.start():].strip()
            return main, unit
    return address.strip(), ""


def split_address(df, address_column):
    main_addresses, unit_numbers = [], []
    for address in df[address_column].astype(str):
        main, unit = find_unit(address)
        main_addresses.append(main)
        unit_numbers.append(unit)
    pos = df.columns.get_loc(address_column) + 1
    df.insert(pos,     'Address Modified', main_addresses)
    df.insert(pos + 1, 'Apt/Unit',         unit_numbers)
    return df


def apply_styles(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for col in ws.iter_cols(1, ws.max_column, 1, 1):
        if col[0].value in ('Address Modified', 'Apt/Unit'):
            col[0].fill = yellow_fill
    wb.save(file_path)


def process_files_in_folder(folder_path):
    for file_name in os.listdir(folder_path):
        if not (file_name.endswith(".xlsx") or file_name.endswith(".csv")):
            continue
        file_path = os.path.join(folder_path, file_name)
        df = pd.read_excel(file_path) if file_name.endswith(".xlsx") else pd.read_csv(file_path)
        if 'ADDRESS' not in df.columns:
            print(f"Skipped {file_name}: no ADDRESS column found.")
            continue
        df = split_address(df, 'ADDRESS')
        if file_name.endswith(".xlsx"):
            df.to_excel(file_path, index=False)
            apply_styles(file_path)
        else:
            df.to_csv(file_path, index=False)
        print(f"Processed: {file_name}")


if __name__ == "__main__":
    folder_path = 'Processed_Files'
    process_files_in_folder(folder_path)
