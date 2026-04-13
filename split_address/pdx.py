import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def split_address(df, address_column):
    addresses = df[address_column].astype(str)
    main_address = []
    unit_numbers = []
    selected_words = ["Apt", "Unit", "#"]  # Words that trigger address modification
    
    for address in addresses:
        match = re.search(r'(?i)\b(Apt|Unit|#)\s*([A-Za-z0-9-]+)', address)
        if match:
            main_address.append(address[:match.start()].strip())
            unit_numbers.append(match.group(0))  # Capture full unit information (e.g., "Apt A11")
        else:
            main_address.append("" if not any(word in address for word in selected_words) else address)
            unit_numbers.append("")
    
    df.insert(df.columns.get_loc(address_column) + 1, 'Address Modified', main_address)
    df.insert(df.columns.get_loc(address_column) + 2, 'Apt/Unit', unit_numbers)
    return df

def apply_styles(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for col in ws.iter_cols():
        if col[0].value in ['Address Modified', 'Apt/Unit']:
            col[0].fill = yellow_fill  # Apply yellow fill only to headers
    
    wb.save(file_path)

def process_files_in_folder(folder_path):
    for file_name in os.listdir(folder_path):
        if file_name.endswith(".xlsx") or file_name.endswith(".csv"):
            file_path = os.path.join(folder_path, file_name)
            if file_name.endswith(".xlsx"):
                df = pd.read_excel(file_path)
            else:
                df = pd.read_csv(file_path)
            
            if 'ADDRESS' in df.columns:
                df = split_address(df, 'ADDRESS')
                
                if file_name.endswith(".xlsx"):
                    df.to_excel(file_path, index=False)
                    apply_styles(file_path)
                else:
                    df.to_csv(file_path, index=False)
            
            print(f"File {file_name} has been processed. New columns 'Address Modified' and 'Apt/Unit' were added.")

# Define folder path
folder_path = 'Processed_Files'
process_files_in_folder(folder_path)
